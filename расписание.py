import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- Конфигурация Смен и Ролей ---
# Роль A: "Утренний стандарт" (Пн-Пт 08-17), Вых: Сб, Вс.
# Роль B: "Вечерний/Выходной" (Ср-Вс: Ср-Пт 11-20, Сб-Вс 09-17), Вых: Пн, Вт.
# Роль C: "Сменный/Закрывающий" (Сб-Вт: Сб-Вс 11-19, Пн-Вт 11-20), Вых: Ср, Чт, Пт.

# Определяем расписание для каждой Роли по дням недели (0=Пн, 6=Вс)
schedules = {
    'A': {
        0: '08:00 - 17:00', 1: '08:00 - 17:00', 2: '08:00 - 17:00', 3: '08:00 - 17:00', 4: '08:00 - 17:00',
        5: 'ВЫХ', 6: 'ВЫХ'
    },
    'B': {
        0: 'ВЫХ', 1: 'ВЫХ',
        2: '11:00 - 20:00', 3: '11:00 - 20:00', 4: '11:00 - 20:00', 5: '09:00 - 17:00', 6: '09:00 - 17:00'
    },
    'C': {
        0: '11:00 - 20:00', 1: '11:00 - 20:00',
        2: 'ВЫХ', 3: 'ВЫХ', 4: 'ВЫХ',
        5: '11:00 - 19:00', 6: '11:00 - 19:00'
    }
}

# Порядок ротации ролей по неделям
# Неделя 1: Сотр1=A, Сотр2=B, Сотр3=C
# Неделя 2: Сотр1=C, Сотр2=A, Сотр3=B (Сдвиг)
# Неделя 3: Сотр1=B, Сотр2=C, Сотр3=A (Сдвиг)
# Неделя 4: Повтор Недели 1 (или продолжение цикла)
rotation_order = ['A', 'B', 'C']

def get_shift(emp_id, day_idx, week_idx):
    # Определяем роль сотрудника на эту неделю
    # Сотрудник 1 (индекс 0) начинает с A (индекс 0 в rotation_order)
    # Формула сдвига: (emp_index + week_index) % 3
    role_idx = (emp_id + week_idx) % 3 
    role = rotation_order[role_idx]
    return schedules[role][day_idx]

# --- Создание Excel ---
wb = Workbook()
ws = wb.active
ws.title = "График работы"

# Стили
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Заголовки дней недели
days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
for col_num, day in enumerate(days, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = day
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    # Установка ширины
    ws.column_dimensions[get_column_letter(col_num)].width = 25

# Генерация календаря (4 недели)
current_row = 2
for week in range(4):
    # Строка с номером недели (опционально, можно пропустить, но для наглядности оставим просто данные)
    # Для высоты ячеек
    ws.row_dimensions[current_row].height = 80 
    
    for day_idx in range(7): # 0..6 (Пн..Вс)
        cell = ws.cell(row=current_row, column=day_idx + 1)
        
        # Формируем текст ячейки
        s1 = get_shift(0, day_idx, week) # Сотрудник 1
        s2 = get_shift(1, day_idx, week) # Сотрудник 2
        s3 = get_shift(2, day_idx, week) # Сотрудник 3
        
        cell_text = (f"Неделя {week+1}\n\n"
                     f"Сотр. 1: {s1}\n"
                     f"Сотр. 2: {s2}\n"
                     f"Сотр. 3: {s3}")
        
        cell.value = cell_text
        cell.alignment = top_left_align
        cell.border = thin_border
        
        # Подсветка выходных дней серым оттенком для наглядности (если это Сб или Вс)
        if day_idx >= 5:
            cell.fill = PatternFill("solid", fgColor="F2F2F2")

    current_row += 1

# Сохранение
file_path = "tabele_kalendar.xlsx"
wb.save(file_path)
print(f"Файл {file_path} успешно создан.")